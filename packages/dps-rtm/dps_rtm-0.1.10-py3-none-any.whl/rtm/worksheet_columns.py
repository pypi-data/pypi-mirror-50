from collections import namedtuple
import openpyxl
from rtm.exceptions import RTMValidatorFileError
from typing import List

WorksheetColumn = namedtuple("WorksheetColumn", "header body index column")


def get_worksheet_columns(path, worksheet_name):
    """Return list of WorksheetColumn objects"""

    # --- Get Workbook ----------------------------------------------------
    wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)

    # --- Get Worksheet ---------------------------------------------------
    ws = None
    for sheetname in wb.sheetnames:
        if sheetname.lower() == worksheet_name.lower():
            ws = wb[sheetname]
    if ws is None:
        raise RTMValidatorFileError(
            f"\nError: Workbook does not contain a '{worksheet_name}' worksheet"
        )

    # --- Convert Worksheet to WorksheetColumn objects --------------------
    ws_data = []
    start_column_num = 1
    for index, col in enumerate(range(start_column_num, ws.max_column + 1)):
        column_header = ws.cell(1, col).value
        column_body = tuple(ws.cell(row, col).value for row in range(2, ws.max_row + 1))
        ws_column = WorksheetColumn(
            header=column_header, body=column_body, index=index, column=col
        )
        ws_data.append(ws_column)

    return ws_data


def get_matching_worksheet_columns(
    sequence_worksheet_columns, field_name: str
) -> List[WorksheetColumn]:
    """Called by constructor to get matching WorksheetColumn objects"""
    matching_worksheet_columns = [
        ws_col
        for ws_col in sequence_worksheet_columns
        if ws_col.header.lower() == field_name.lower()
    ]
    return matching_worksheet_columns


def get_first_matching_worksheet_column(
    sequence_worksheet_columns, field_name: str
) -> WorksheetColumn:
    cols = get_matching_worksheet_columns(sequence_worksheet_columns, field_name)
    return cols[0]
