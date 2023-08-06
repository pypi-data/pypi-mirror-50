import json
from functools import reduce
from operator import getitem

from openpyxl import Workbook

from jsondler.json_tools.requests import tuplize_json_coord_pairs
from jsondler.json_tools.serialization import serialize_field
from xlson._lib.coords_tools import get_xl_coord
from xlson.scheme import EntityInfo, XLSonScheme, XLSonSheetScheme
from xlson.formatting import CELL_DEFAULT_META, fill_cell_meta


class XLSonHandler(XLSonScheme):

    @property
    def XLSonSheetHandler(self):
        return XLSonSheetHandler

    @classmethod
    def load(cls, xlson_path):
        with open(xlson_path) as xlson_f:
            xlson = tuplize_coords_in_prep_xl(json.load(xlson_f))
        return cls.load_from_dict(in_dict=xlson)

    def dump(self, xlson_path):
        xlson_f = open(xlson_path, 'w')
        json.dump(self.xlson, xlson_f, indent=2, default=serialize_field, ensure_ascii=False)
        xlson_f.close()
        return xlson_path

    def to_xlsx(self, xlsx_path):
        wb = Workbook()
        ws_0 = wb.active
        for xlson_sheet in [self.main_sheet] + \
                           [self.XLSonSheetHandler.load_from_dict(supp_sheet) for supp_sheet in self.supp_sheets]:
            ws = wb.create_sheet(xlson_sheet.sheet_name)
            data_df = xlson_sheet.data_df
            meta_df = xlson_sheet.meta_df
            for i in range(len(data_df)):
                for j in range(len(data_df[0])):
                    xl_coord = None
                    xl_coord = get_xl_coord(i, j)
                    ws[xl_coord] = data_df[i][j]
                    fill_cell_meta(ws[xl_coord], meta_df[i][j])
                    if meta_df[i][j]["merged_with"]:
                        ws.merge_cells(xl_coord+":"+get_xl_coord(*meta_df[i][j]["merged_with"][-1]))
        wb.remove_sheet(ws_0)
        wb.save(xlsx_path)

    @property
    def xlson(self):
        return XLSonScheme(main_sheet=self.main_sheet.get_json(),
                           supp_sheets=self.supp_sheets,
                           source_path=self.source_path).get_json()

    def get_json(self):
        return self.xlson

    @classmethod
    def load_from_dict(cls, in_dict):
        xlson_handler = super(XLSonHandler, cls).load_from_dict(in_dict=in_dict)
        xlson_handler.main_sheet = XLSonSheetHandler.load_from_dict(xlson_handler.main_sheet, main_sheet=True)
        return xlson_handler

    def get_supp_sheet(self, ):
        return

    def iterate_supp_sheets(self):
        yield

    def __getitem__(self, item):
        return self.__dict__[item]

    def __setitem__(self, key, value):
        self.__dict__[key] = value

    def __eq__(self, other):
        return self.xlson == other.xlson


class XLSonSheetHandler(XLSonSheetScheme):

    class MainSheetScheme(XLSonSheetScheme):
        main_sheet = True

    def get_json(self):
        # Hardcode keys - be careful
        got_json = super(XLSonSheetHandler, self).get_json()
        meta_df_path = XLSonSheetScheme.attr_scheme()["meta_df"]
        reduce(getitem, meta_df_path[:-1], got_json)[meta_df_path[-1]] = self.meta_df.meta_df
        return got_json

    @classmethod
    def load_from_dict(cls, in_dict, main_sheet=False, cell_default_meta=None, additional_entity_attrs=None):
        if main_sheet:
            sheet_scheme = cls.MainSheetScheme.load_from_dict(in_dict=in_dict)
        else:
            sheet_scheme = super(XLSonSheetHandler, cls).load_from_dict(in_dict=in_dict)
        sheet_handler = cls.load_from_scheme(sheet_scheme)
        sheet_handler.main_sheet = main_sheet

        if cell_default_meta is not None:
            sheet_handler.cell_default_meta = cell_default_meta
        elif not sheet_handler.cell_default_meta:
            sheet_handler.cell_default_meta = CELL_DEFAULT_META
        sheet_handler.meta_df = CellsMetaDF(sheet_handler.meta_df, sheet_handler.cell_default_meta)

        if additional_entity_attrs is None:
            additional_entity_attrs = cls.get_additional_entity_attrs(sheet_handler.entities)
        sheet_handler.additional_entities_attrs = additional_entity_attrs

        return sheet_handler

    @classmethod
    def load_from_scheme(cls, sheet_scheme):
        sheet_handler = cls()
        sheet_handler.__dict__ = sheet_scheme.__dict__
        sheet_handler.main_sheet = sheet_scheme.main_sheet
        return sheet_handler

    def get_entity(self, ):
        return

    def iterate_entities(self):
        yield

    def __getitem__(self, item):
        return self.__dict__[item]

    def __setitem__(self, key, value):
        self.__dict__[key] = value

    @property
    def sheet_name(self):
        return self._sheet_name

    @sheet_name.setter
    def sheet_name(self, value):
        # TODO: write code for managing source dict
        self._sheet_name = value

    @staticmethod
    def get_additional_entity_attrs(entities_list):
        """

        :param entities_list: list of
        :return:
        """
        additional_entity_attrs = list()
        for entity_dict in entities_list:
            for attr in entity_dict.keys():
                if attr not in EntityInfo.attr_scheme() and attr not in additional_entity_attrs:
                    additional_entity_attrs.append(attr)
        return additional_entity_attrs

    @property
    def EntityHandler(self):
        class EntityHandler(EntityInfo):
            additional_attrs = self.additional_entities_attrs

        return EntityHandler

    def get_EntityHandler(self, additional_attrs=None):
        if additional_attrs is not None:
            self.additional_entities_attrs = additional_attrs
        return self.EntityHandler


class CellsMetaDF(object):

    def __init__(self, meta_df, cell_default_meta):
        self.meta_df = meta_df
        self.cell_default_meta = cell_default_meta

    def __getitem__(self, item):
        if item < 0:
            item = len(self.meta_df)-item
        return CellsMetaRow(self.meta_df[item], self.cell_default_meta)

    def __setitem__(self, key, value):
        self.meta_df[key] = value

    def __eq__(self, other):
        return self.meta_df == other.meta_df

    def __len__(self):
        return len(self.meta_df)


class CellsMetaRow(object):

    def __init__(self, meta_row, cell_default_meta):
        self.meta_row = meta_row
        self.cell_default_meta = cell_default_meta

    def __getitem__(self, item):
        if item < 0:
            item = len(self.meta_row)-item
        if self.meta_row[item]:
            # return deepupdate(deepcopy(self.cell_default_meta), self.meta_row[item])
            # return CellMeta(cell_default_meta=self.cell_default_meta, cell_diff_meta=self.meta_row[item])
            return self.meta_row[item]
        else:
            return self.cell_default_meta

    def __setitem__(self, key, value):
        if value:
            # self.meta_row[key] = deepdiff(self.cell_default_meta, value)
            self.meta_row[key] = value
        else:
            # self.meta_row[key] = {}
            self.meta_row[key] = None

    def __len__(self):
        return len(self.meta_row)


class CellMeta(object):

    def __init__(self, cell_default_meta, cell_diff_meta):
        self.cell_default_meta = cell_default_meta
        self.cell_diff_meta = cell_diff_meta

    def __getitem__(self, item):
        return self.cell_diff_meta.get(item, self.cell_default_meta[item])


def tuplize_coords_in_prep_xl(prep_xl_dict):
    coords_paths = [
        ("main_sheet", "meta_df", "*", "*", "merged_to"),
        ("main_sheet", "meta_df", "*", "*", "merged_with", "*"),
        ("supp_sheets", "*", "meta_df", "*", "*", "merged_to"),
        ("supp_sheets", "*", "meta_df", "*", "*", "merged_with", "*"),
    ]

    return tuplize_json_coord_pairs(prep_xl_dict, *coords_paths)
