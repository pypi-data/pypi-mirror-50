import datetime
from copy import deepcopy

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple
from xlrd import open_workbook, xldate_as_tuple

from xlson.constants import xlson_logger
from xlson.formatting import cell_meta_to_dict
from xlson.handlers import XLSonHandler, XLSonSheetHandler
from xlson.formatting import CELL_DEFAULT_META
from xlson._lib.general_utils import digitalize_str
from xlson._lib.coords_tools import coords_from_range


def prepare_xl(xl_path, data_only=False, values_strip=None, digitalization=True, crop_empty=True, n_rows=None):
    try:
        return prepare_new_xl(new_xl_path=xl_path,
                              data_only=data_only,
                              values_strip=values_strip,
                              digitalization=digitalization,
                              crop_empty=crop_empty,
                              n_rows=n_rows)
    # except InvalidFileException:
    except:
        xlson_logger.warning("%s cannot be prepared as new Excel format - trying old Excel format preparation" % xl_path)
        try:
            return prepare_old_xl(old_xl_path=xl_path,
                                  values_strip=values_strip,
                                  digitalization=digitalization,
                                  crop_empty=crop_empty,
                                  n_rows=n_rows)
        except:
            xlson_logger.warning("cannot read '%s'" % xl_path)
            return XLSonHandler(
                main_sheet=XLSonSheetHandler.load_from_dict({
                    "data_df": [["an error", "occurred", "while", "processing", xl_path]],
                    "meta_df": [[None] * 5],
                    "entities": XLSonSheetHandler.entites_0,
                    "supp_sheets": XLSonSheetHandler.supp_sheets_0,
                    "cell_default_meta": CELL_DEFAULT_META,
                }, main_sheet=True),
                supp_sheets = list(),
                source_path = xl_path,
            )


def prepare_new_xl(new_xl_path, data_only=False, values_strip=None, digitalization=True, crop_empty=True, n_rows=None):
    xlson_logger.info("%s conversion to xlson started" % new_xl_path)
    main_sheet = dict()
    supp_sheets_list = list()
    wb = load_workbook(new_xl_path, data_only=data_only)
    n = 0
    for sheet_name in wb.sheetnames:
        merged_cells_dict = get_merged_cells(wb[sheet_name])
        sheet_dict = {
            'cell_default_meta': CELL_DEFAULT_META,
            'sheet_name': sheet_name,
            'data_df': iterate_sheet(wb[sheet_name],
                                     cell_func=get_cell_value,
                                     add_args_dict={'value_strip': values_strip,
                                                    'digitalization': digitalization},
                                     n_rows=n_rows),
            'entities': deepcopy(XLSonSheetHandler.entites_0),
        }

        if crop_empty:
            last_cell = get_last_cell(sheet_dict['data_df'])
            sheet_dict['data_df'] = [sheet_dict['data_df'][i][:last_cell[1]+1] for i in range(last_cell[0]+1)]
            sheet_dict['meta_df'] = iterate_sheet(wb[sheet_name],
                                                  cell_func=cell_meta_to_dict,
                                                  add_args_dict={'merged_cells_dict': merged_cells_dict},
                                                  last_cell=last_cell,
                                                  n_rows=n_rows)
        else:
            sheet_dict['meta_df'] = iterate_sheet(wb[sheet_name],
                                                  cell_func=cell_meta_to_dict,
                                                  add_args_dict={'merged_cells_dict': merged_cells_dict},
                                                  n_rows=n_rows)

        if n == 0:
            main_sheet = sheet_dict
        else:
            supp_sheets_list.append(sheet_dict)
        n += 1
    main_sheet['supp_sheets'] = [supp_sheet['sheet_name'] for supp_sheet in supp_sheets_list]
    xlson_logger.info("%s conversion to xlson finished" % new_xl_path)
    return XLSonHandler(main_sheet=XLSonSheetHandler.load_from_dict(main_sheet, main_sheet=True),
                        supp_sheets=supp_sheets_list,
                        source_path=new_xl_path)


def get_last_cell(data_df):
    max_row = 0
    max_col = 0
    for i in range(len(data_df)):
        for j in range(len(data_df[i])):
            if data_df[i][j] is not None:
                if i > max_row:
                    max_row = i
                if j > max_col:
                    max_col = j
    return max_row, max_col


def get_merged_cells(sheet):
    merged_cells_dict = {
        "merged_with": {},
        "merged_to": {},
    }

    mc_ranges = sheet.merged_cells.ranges
    for mc_range in mc_ranges:
        c_list = mc_range.coord.split(":")
        first_c = tuple(map(lambda c: c - 1, coordinate_to_tuple(c_list[0])))
        last_c = tuple(map(lambda c: c - 1, coordinate_to_tuple(c_list[1])))
        merged_coords_list = coords_from_range(first_c, last_c)
        merged_cells_dict["merged_with"][first_c] = merged_coords_list[1:]
        for merged_coord in merged_coords_list[1:]:
            merged_cells_dict["merged_to"][merged_coord] = first_c
    return merged_cells_dict


def iterate_sheet(sheet, cell_func=None, add_args_dict=None, last_cell=None, n_rows=None):
    rows_list = list()
    i = 0
    for row in sheet:
        if last_cell is not None and i > last_cell[0]:
            break
        if type(n_rows) is int and i >= n_rows:
            break
        curr_row_list = list()
        j = 0
        for cell in row:
            if last_cell is not None and j > last_cell[1]:
                break
            if callable(cell_func):
                if type(add_args_dict) is dict:
                    curr_row_list.append(cell_func(cell, **add_args_dict))
                elif type(add_args_dict) is list or type(add_args_dict) is tuple:
                    curr_row_list.append(cell_func(cell, *add_args_dict))
                else:
                    curr_row_list.append(cell_func(cell))
            else:
                curr_row_list.append(cell)
            j += 1
        rows_list.append(curr_row_list)
        i += 1
    return rows_list


def get_cell_value(cell, value_strip=None, digitalization=True, special_formating=None, **kwargs):
    if callable(special_formating):
        v = special_formating(cell, **kwargs)
    else:
        v = cell.value
    if type(v) is datetime.datetime:
        return v.strftime("%d.%m.%Y")
    if type(v) is str:
        if not v:
            return None
        if type(value_strip) is str or value_strip is None:
            if digitalization:
                return digitalize_str(v.strip(value_strip))
            else:
                return v.strip(value_strip)
    return v


def prepare_old_xl(old_xl_path, values_strip=None, digitalization=True, crop_empty=True, n_rows=None):
    # TODO: implement formatting info conversion to meta_df
    xlson_logger.info("%s conversion to xlson started" % old_xl_path)
    main_sheet = dict()
    supp_sheets_list = list()
    wb = open_workbook(old_xl_path, formatting_info=True)  # data_only equiv has not been found
    n = 0
    for sheet_name in wb.sheet_names():
        # merged_cells_dict = get_merged_cells(wb.sheet_by_name(sheet_name))  # TODO: implement meged cells preparation
        sheet_dict = {
            'cell_default_meta': CELL_DEFAULT_META,
            'sheet_name': sheet_name,
            'data_df': iterate_sheet(wb.sheet_by_name(sheet_name).get_rows(),
                                     cell_func=get_cell_value,
                                     add_args_dict={'value_strip': values_strip,
                                                    'digitalization': digitalization,
                                                    'special_formating': _check_xlrd_types,
                                                    'datemode': wb.datemode},
                                     n_rows=n_rows),
            'entities': deepcopy(XLSonSheetHandler.entites_0),
        }

        if crop_empty:
            last_cell = get_last_cell(sheet_dict['data_df'])
            if last_cell == (0, 0):
                sheet_dict['data_df'] = [[None]]
            else:
                sheet_dict['data_df'] = [sheet_dict['data_df'][i][:last_cell[1]+1] for i in range(last_cell[0]+1)]
            sheet_dict['meta_df'] = [[None] * (last_cell[1]+1)] * (last_cell[0]+1)
            # TODO: fill meta_df
            # sheet_dict['meta_df'] = iterate_sheet(wb[sheet_name],
            #                                       cell_func=cell_meta_to_dict,
            #                                       add_args_dict={'merged_cells_dict': merged_cells_dict},
            #                                       last_cell=last_cell,
            #                                       n_rows=n_rows)
        else:
            sheet_dict['meta_df'] = [[None] * wb.sheet_by_name(sheet_name).ncols] * wb.sheet_by_name(sheet_name).nrows
            # TODO: fill meta_df
            # sheet_dict['meta_df'] = iterate_sheet(wb[sheet_name],
            #                                       cell_func=cell_meta_to_dict,
            #                                       add_args_dict={'merged_cells_dict': merged_cells_dict},
            #                                       n_rows=n_rows)

        if n == 0:
            main_sheet = sheet_dict
        else:
            supp_sheets_list.append(sheet_dict)
        n += 1
    main_sheet['supp_sheets'] = [supp_sheet['sheet_name'] for supp_sheet in supp_sheets_list]
    xlson_logger.info("%s conversion to xlson finished" % old_xl_path)
    return XLSonHandler(main_sheet=XLSonSheetHandler.load_from_dict(main_sheet, main_sheet=True),
                        supp_sheets=supp_sheets_list,
                        source_path=old_xl_path)


def _check_xlrd_types(cell, **kwargs):
    v = cell.value
    if cell.ctype == 0 or cell.ctype == 6:
        return None
    if cell.ctype == 2:
        if v - float(int(v)) > 0.0:
            return v
        else:
            return int(v)
    if cell.ctype == 3:
        return datetime.datetime(*xldate_as_tuple(v, kwargs.get("datemode", 0)))
        # return datetime.datetime(1900, 1, 1) + datetime.timedelta(int(v)-2)
    return v
